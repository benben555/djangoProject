from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl import load_workbook

from webapp import models
from webapp.utils.pagination import Pagination


def depart_list(request):
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset, page_size=10)
    context = {"queryset": page_object.page_queryset,
               "page_string": page_object.html()}
    return render(request, "depart_list.html", context)


def depart_add(request):
    if request.method == "GET":
        return render(request, "depart_add.html")
    title = request.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")


def depart_delete(request):
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(request, nid):
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {"row_object": row_object})
    title = request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")


def depart_multi(request):
    # 1获取用户上传文件对象
    file_object = request.FILES.get('exc')
    # 2对象传递给openpyxl
    # work_book_object = load_workbook("文件路径")
    work_book_object = load_workbook(file_object)
    sheet = work_book_object.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        models.Department.objects.create(title=text)
    return HttpResponse("ok")
